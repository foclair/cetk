"""Command line interface for managing a Clair emission inventory offline."""

import argparse
import datetime
import os
import sys
from math import ceil
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

import etk
from etk import logging
from etk.db import run_migrate
from etk.tools.utils import (
    CalledProcessError,
    SubprocessError,
    add_standard_command_options,
    check_and_get_path,
    create_from_template,
    get_db,
    get_template_db,
)

log = logging.getLogger("etk")

settings = etk.configure()

from etk.edb.const import DEFAULT_SRID, SHEET_NAMES  # noqa
from etk.edb.exporters import export_sources  # noqa
from etk.edb.importers import (  # noqa
    import_activitycodesheet,
    import_codesetsheet,
    import_gridsources,
    import_sourceactivities,
    import_sources,
    import_timevarsheet,
)
from etk.edb.models import Settings, Substance  # noqa
from etk.edb.rasterize.rasterizer import EmissionRasterizer, Output  # noqa
from etk.emissions.calc import aggregate_emissions, get_used_substances  # noqa
from etk.emissions.views import (  # noqa
    create_areasource_emis_table,
    create_pointsource_emis_table,
)

SOURCETYPES = ("point", "area", "grid")
DEFAULT_EMISSION_UNIT = "kg/year"

sheet_choices = ["All"]
sheet_choices.extend(SHEET_NAMES)


class DryrunAbort(Exception):
    """Forcing abort of database changes when doing dryrun."""

    pass


def adjust_extent(extent, srid, cellsize):
    """adjust extent to include an integer nr of cells."""
    x1, y1, x2, y2 = (
        extent or Settings.get_current().extent.transform(srid, clone=True).extent
    )
    nx = ceil((x2 - x1) / cellsize)
    ny = ceil((y2 - y1) / cellsize)
    x2 = x1 + nx * cellsize
    y2 = y1 + ny * cellsize
    extent = (x1, y1, x2, y2)
    # Settings.extent is a Polygon, Settings.extent.extent a tuple (x1, y1, x2, y2)
    if extent is None:
        log.error("could not rasterize emissions, default extent not set")
        sys.exit(1)
    return extent, ny, nx


class Editor(object):
    def __init__(self):
        self.db_path = settings.DATABASES["default"]["NAME"]

    def migrate(self, template=False, db_path=None):
        if db_path is None:
            if template:
                db_path = get_template_db()
            else:
                db_path = get_db()

        log.debug(f"Running migrations for database {db_path}")
        try:
            std_out, std_err = run_migrate(db_path=db_path)
        except (CalledProcessError, SubprocessError) as err:
            log.error(f"Error while migrating {db_path}: {err}")
        log.debug(f"Successfully migrated database {db_path}")

    def import_workbook(self, filename, sheets=SHEET_NAMES, dry_run=False):
        return_msg = []
        db_updates = {}
        workbook = load_workbook(filename=filename, data_only=True)
        import_sheets = [
            s
            for s in SHEET_NAMES
            if s in frozenset(sheets) and s in frozenset(workbook.sheetnames)
        ]
        try:
            with transaction.atomic():
                for sheet in import_sheets:
                    if sheet not in workbook.sheetnames:
                        log.info(f"Workbook has no sheet named {sheet}, skipped import")
                if any(name != "GridSource" for name in import_sheets):
                    updates, msgs = import_sourceactivities(
                        filename,
                        import_sheets=import_sheets,
                        validation=dry_run,
                    )
                db_updates.update(updates)
                if len(msgs) != 0:
                    return_msg += msgs
                    if not dry_run:
                        raise ImportError(return_msg)
                if dry_run:
                    raise DryrunAbort
            # grid sources imported outside atomic transaction
            # to avoid errors when writing rasters using rasterio
            if "GridSource" in import_sheets:
                updates, msgs = import_gridsources(filename, validation=dry_run)
                db_updates.update(updates)
                if len(msgs) > 0:
                    return_msg += msgs
                    raise ImportError(return_msg)
        except DryrunAbort:
            # validate grid sources
            if "GridSource" in import_sheets:
                updates, msgs = import_gridsources(filename, validation=True)
            if len(msgs) != 0:
                return_msg += msgs
                log.error(
                    f"Errors during import:{os.linesep}{os.linesep.join(return_msg)}"
                )
                log.info("Finished dry-run")
            else:
                log.info("Successful dry-run, no errors, the file is ready to import.")
        except ImportError:
            log.error(f"Errors during import:{os.linesep}{os.linesep.join(return_msg)}")
        else:
            log.info(
                datetime.datetime.now().strftime("%H:%M:%S")
                + f" imported data {db_updates}"
            )
        return db_updates, return_msg

    def update_emission_tables(
        self, sourcetypes=None, unit=DEFAULT_EMISSION_UNIT, substances=None
    ):
        sourcetypes = sourcetypes or SOURCETYPES
        substances = substances or get_used_substances()
        if "point" in sourcetypes:
            create_pointsource_emis_table(substances=substances, unit=unit)
        if "area" in sourcetypes:
            create_areasource_emis_table(substances=substances, unit=unit)

    def aggregate_emissions(
        self,
        filename,
        sourcetypes=None,
        unit=DEFAULT_EMISSION_UNIT,
        codeset=None,
        substances=None,
    ):
        df = aggregate_emissions(
            sourcetypes=sourcetypes, unit=unit, codeset=codeset, substances=substances
        )
        try:
            df.to_csv(filename, sep=";")
        except Exception as err:
            log.error(
                f"could not write aggregated emission to file {filename}: {str(err)}"
            )
            sys.exit(1)

    def rasterize_emissions(
        self,
        outputpath,
        cellsize,
        sourcetypes=None,
        unit=DEFAULT_EMISSION_UNIT,
        codeset=None,
        substances=None,
        begin=None,
        end=None,
        extent=None,
        srid=None,
        timezone=None,
    ):
        substances = substances or get_used_substances()
        timezone = timezone or datetime.timezone.utc
        srid = srid or DEFAULT_SRID
        extent, ny, nx = adjust_extent(extent, srid, cellsize)
        try:
            output = Output(
                extent=extent, timezone=timezone, path=outputpath, srid=srid
            )
            rasterizer = EmissionRasterizer(output, nx=nx, ny=ny)
            rasterizer.process(substances, begin=begin, end=end, unit=unit)
        except Exception as err:
            log.error(f"could not rasterize emissions: {str(err)}")
            sys.exit(1)

    def export_data(self, filename):
        export_sources(filename)
        return True


def main():
    db_path = get_db() or "unspecified"
    parser = argparse.ArgumentParser(
        description="Manage Clair offline emission inventories",
        usage=f"""etk <command> [<args>]

        Main commands are:
        create   create an sqlite inventory
        migrate  migrate an sqlite inventory
        import   import data
        export   export data
        calc     calculate emissions

        Current database is {db_path} (set by $ETK_DATABASE_PATH)
        """,
    )
    add_standard_command_options(parser)
    parser.add_argument(
        "command",
        help="Subcommand to run",
        choices=("migrate", "create", "import", "export", "calc"),
    )
    verbosity = [arg for arg in sys.argv if arg == "-v"]
    sys_args = [arg for arg in sys.argv if arg != "-v"]
    sub_args = sys_args[2:]
    main_args = parser.parse_args(args=sys_args[1:2] + verbosity)
    editor = Editor()
    if main_args.command == "create":
        sub_parser = argparse.ArgumentParser(
            description="Create database from template.",
            usage="etk create <filename>",
        )
        sub_parser.add_argument(
            "filename",
            help="Path of new database",
        )
        args = sub_parser.parse_args(sub_args)
        create_from_template(args.filename)
        log.debug(
            "Created new database '{args.filename}' from template '{get_template_db()}'"
        )
        sys.exit(0)

    if (
        len(sys.argv) < 2
        and sys.argv[2] not in ("-h", "--help", "--version")
        and db_path == "unspecified"
    ):
        sys.stderr.write("No database specified, set by $ETK_DATABASE_PATH\n")
        sys.exit(1)

    if main_args.command == "migrate":
        sub_parser = argparse.ArgumentParser(
            description=f"Migrate database {db_path}.",
            usage="usage: etk migrate",
        )
        sub_parser.add_argument(
            "--template",
            action="store_true",
            help="Migrate the template database",
        )
        sub_parser.add_argument(
            "--dbpath",
            help="Specify database path manually",
        )
        args = sub_parser.parse_args(sub_args)
        editor.migrate(template=args.template, db_path=args.dbpath)
    elif main_args.command == "import":
        sub_parser = argparse.ArgumentParser(
            description="Import data from an xlsx-file",
            usage="etk import <filename> [options]",
        )
        sub_parser.add_argument(
            "filename", help="Path to xslx-file", type=check_and_get_path
        )
        sub_parser.add_argument(
            "--sheets",
            nargs="+",
            default=SHEET_NAMES,
            help=f"List of sheets to import, valid names {SHEET_NAMES}",
        )
        sub_parser.add_argument(
            "--dryrun",
            action="store_true",
            help="Do dry run to validate import file without actually importing data",
        )
        sub_parser.add_argument(
            "--residential-heating",
            action="store_true",
            help="Import file with energy demand for residential heating",
        )
        sub_parser.add_argument(
            "--substances",
            nargs="*",
            help="Only import residential heating emissions for these substances"
            + " (default is all with emissions)",
            choices=Substance.objects.values_list("slug", flat=True),
            metavar=("NOx", "PM10"),
        )
        # pointsource_grp = sub_parser.add_argument_group(
        #     "pointsources", description="Options for pointsource import"
        # )
        args = sub_parser.parse_args(sub_args)
        if not Path(db_path).exists():
            sys.stderr.write(
                "Database " + db_path + " does not exist, first run "
                "'etk create' or 'etk migrate'\n"
            )
            sys.exit(1)

        editor.import_workbook(args.filename, sheets=args.sheets, dry_run=args.dryrun)
        sys.exit(0)
    elif main_args.command == "calc":
        sub_parser = argparse.ArgumentParser(
            description="Calculate emissions",
            usage="etk calc [options]",
        )
        sub_parser.add_argument(
            "--unit",
            default=DEFAULT_EMISSION_UNIT,
            help="Unit of emissions, default=%(default)s",
        )
        sub_parser.add_argument(
            "--sourcetypes", nargs="*", help="Only sourcetypes", choices=SOURCETYPES
        )
        sub_parser.add_argument(
            "--substances",
            nargs="*",
            help="Only substances (default is all with emissions)",
            choices=Substance.objects.values_list("slug", flat=True),
            metavar=("NOx", "PM10"),
        )
        calc_grp = sub_parser.add_mutually_exclusive_group()
        calc_grp.add_argument(
            "--update", help="Create/update emission tables", action="store_true"
        )
        calc_grp.add_argument(
            "--aggregate", help="Aggregate emissions", metavar="FILENAME"
        )
        aggregate_grp = sub_parser.add_argument_group(
            "aggregate emissions", description="Options to aggregate emissions"
        )
        aggregate_grp.add_argument(
            "--codeset", help="Aggregate emissions by codeset", metavar="SLUG"
        )
        calc_grp.add_argument(
            "--rasterize", help="Rasterize emissions", metavar="OUTPUTPATH"
        )
        rasterize_grp = sub_parser.add_argument_group(
            "rasterize emissions", description="Settings to rasterize emissions"
        )
        rasterize_grp.add_argument(
            "--cellsize", help="Cellsize (meter) in output raster", type=float
        )
        rasterize_grp.add_argument(
            "--extent",
            help="Extent of output raster. Settings.extent is taken otherwise",
            nargs=4,
            type=float,
            metavar=("x1", "y1", "x2", "y2"),
        )
        rasterize_grp.add_argument(
            "--srid",
            help="EPSG of output raster. 4-5 digits integer",
            metavar="EPSG",
        )
        rasterize_grp.add_argument(
            "--begin",
            help="when hourly rasters are desired, specify begin date."
            + " Time 00:00 assumed",
            metavar="2022-01-01",
        )
        rasterize_grp.add_argument(
            "--end",
            help="when hourly rasters are desired, specify end date"
            + " Time 00:00 assumed",
            metavar="2023-01-01",
        )
        # TODO add argument begin/end for rasterize
        # TODO add argument to aggregate emissions within polygon

        args = sub_parser.parse_args(sub_args)
        try:
            if args.substances is not None:
                substances = []
                for s in args.substances:
                    substances.append(Substance.objects.get(slug=s))
            else:
                substances = None
        except Substance.DoesNotExist:
            sys.stderr.write(f"Substance {s} does not exist.\n")
        if not Path(db_path).exists():
            sys.stderr.write("Database does not exist.\n")
            sys.exit(1)
        if args.update:
            editor.update_emission_tables(sourcetypes=args.sourcetypes, unit=args.unit)
            sys.stdout.write("Successfully updated tables\n")
            sys.exit(0)
        if args.aggregate is not None:
            editor.aggregate_emissions(
                args.aggregate,
                substances=substances,
                sourcetypes=args.sourcetypes,
                unit=args.unit,
                codeset=args.codeset,
            )
            sys.stdout.write("Successfully aggregated emissions\n")
            sys.exit(0)
        if args.rasterize is not None:
            if args.begin is not None:
                args.begin = datetime.datetime.strptime(args.begin, "%Y-%m-%d").replace(
                    tzinfo=datetime.timezone.utc
                )
                if args.end is not None:
                    args.end = datetime.datetime.strptime(args.end, "%Y-%m-%d").replace(
                        tzinfo=datetime.timezone.utc
                    )
                else:
                    sys.stderr.write(
                        "If begin is specified," + " end has to be specified too.\n"
                    )
                    sys.exit(1)
            editor.rasterize_emissions(
                args.rasterize,
                args.cellsize,
                substances=substances,
                sourcetypes=args.sourcetypes,
                unit=args.unit,
                extent=args.extent,
                srid=args.srid,
                begin=args.begin,
                end=args.end,
            )  # TODO add arguments for codeset, substances, begin/end, timezone!
            # could also add for polygon, but this filtering is not implemented yet!!
            sys.stdout.write("Successfully rasterized emissions\n")
            sys.exit(0)

    elif main_args.command == "export":
        sub_parser = argparse.ArgumentParser(
            description="Export data to xlsx-file",
            usage="etk export <filename> [options]",
        )
        sub_parser.add_argument("filename", help="Path to xslx-file")
        if not Path(db_path).exists():
            sys.stderr.write(
                "Database " + db_path + " does not exist, first run "
                "'etk create' or 'etk migrate'\n"
            )
            sys.exit(1)

        args = sub_parser.parse_args(sub_args)
        try:
            # Check if the file can be created at the given path
            os.access(args.filename, os.W_OK | os.X_OK)
        except Exception as e:
            sys.stderr.write(f"File {args.filename} cannot be created: \n {e} ")
            sys.exit(1)
        status = editor.export_data(args.filename)
        if status:
            sys.stdout.write(f"Exported data from {db_path} to {args.filename}.\n")
            sys.exit(0)
        else:
            sys.stderr.write("Did not export data, something went wrong.")
            sys.exit(1)
