"""Data importers for the edb application."""

import logging
from itertools import islice

import numpy as np
import pandas as pd
from django.contrib.gis.geos import Point
from django.db import IntegrityError
from openpyxl import load_workbook

from etk.edb.const import SHEET_NAMES, WGS84_SRID
from etk.edb.models.common_models import Settings
from etk.edb.models.eea_emfacs import EEAEmissionFactor
from etk.edb.models.source_models import (
    Activity,
    ActivityCode,
    CodeSet,
    EmissionFactor,
    Facility,
    PointSource,
    PointSourceActivity,
    PointSourceSubstance,
    Substance,
    Timevar,
    VerticalDist,
)
from etk.edb.units import (
    activity_ef_unit_to_si,
    activity_rate_unit_to_si,
    emission_unit_to_si,
)
from etk.tools.utils import cache_queryset

# column facility and name are used as index and is therefore not included here
REQUIRED_COLUMNS = {
    "facility_id": np.str_,
    "lat": float,
    "lon": float,
    "facility_name": np.str_,
    "source_name": np.str_,
    "timevar": np.str_,
    # "activitycode1": np.str_,
    # "activitycode2": np.str_,
    # "activitycode3": np.str_,
    "chimney_height": float,
    "outer_diameter": float,
    "inner_diameter": float,
    "gas_speed": float,
    "gas_temperature": float,
    "house_width": float,
    "house_height": float,
}


log = logging.getLogger(__name__)


class ImportError(Exception):
    """Error while importing emission data."""

    pass


def import_error(message, return_message="", dry_run=False):
    if not dry_run:
        raise ImportError(message)
    else:
        return_message += "\n" + message + "\n"
    return return_message


def cache_pointsources(queryset):
    """Return dict of model instances with (facility__official_id, name): instance"""
    sources = {}
    for source in queryset:
        if source.facility is not None:
            sources[source.facility.official_id, source.name] = source
        else:
            sources[None, source.name] = source
    return sources


def cache_codeset(code_set):
    if code_set is None:
        return {}
    return cache_queryset(code_set.codes.all(), "code")


def worksheet_to_dataframe(data):
    cols = next(data)
    data = list(data)
    data = (islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns=cols)
    # remove empty rows
    empty_count = 0
    for ind in range(-1, -1 * len(df), -1):
        if all([pd.isnull(val) for val in df.iloc[ind]]):
            empty_count += 1
        else:
            break
    # remove the last 'empty_count' lines
    df = df.head(df.shape[0] - empty_count)
    # remove empty columns without label
    if None in df.columns:
        if np.all([pd.isnull(val) for val in df[None].values]):
            df = df.drop(columns=[None])
    return df


def import_pointsources(
    filepath, return_message="", dry_run=False, encoding=None, srid=None
):
    """Import point-sources from xlsx or csv-file.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
        srid: srid of file, default is same srid as domain
    """
    # or change to user defined SRID?
    try:
        project_srid = Settings.objects.get().srid
    except Settings.DoesNotExist:
        project_srid = WGS84_SRID
    # cache related models
    substances = cache_queryset(Substance.objects.all(), "slug")
    timevars = cache_queryset(Timevar.objects.all(), "name")
    facilities = cache_queryset(Facility.objects.all(), "official_id")
    pointsources = cache_pointsources(
        PointSource.objects.select_related("facility")
        .prefetch_related("substances")
        .all()
    )

    # using filter.first() here, not get() because code_set{i} does not have to exist
    code_sets = [
        cache_codeset(CodeSet.objects.filter(id=i).first()) for i in range(1, 4)
    ]

    extension = filepath.suffix
    if extension == ".csv":
        # read csv-file
        with open(filepath, encoding=encoding or "utf-8") as csvfile:
            log.debug("reading point-sources from csv-file")
            df = pd.read_csv(
                csvfile,
                sep=";",
                skip_blank_lines=True,
                comment="#",
                dtype=REQUIRED_COLUMNS,
            )
    elif extension == ".xlsx":
        # read spreadsheet
        try:
            workbook = load_workbook(filename=filepath, data_only=True)
        except Exception as exc:
            return_message = import_error(str(exc), return_message, dry_run)
        worksheet = workbook.worksheets[0]
        if len(workbook.worksheets) > 1:
            log.debug("Multiple sheets in spreadsheet, importing sheet 'PointSource'.")
            data = workbook["PointSource"].values
        else:
            data = worksheet.values
        df = worksheet_to_dataframe(data)
        df = df.astype(dtype=REQUIRED_COLUMNS)
        # below is necessary not to create facilities with name 'None'
        df = df.replace(to_replace="None", value=None)
    else:
        return_message = import_error(
            "Only xlsx and csv files are supported for import", return_message, dry_run
        )
    for col in REQUIRED_COLUMNS.keys():
        if col not in df.columns:
            return_message = import_error(
                f"Missing required column '{col}'", return_message, dry_run
            )

    # set dataframe index
    try:
        df.set_index(
            ["facility_id", "source_name"], verify_integrity=True, inplace=True
        )
    except ValueError as err:
        return_message = import_error(
            f"Non-unique combination of facility_id and source_name: {err}",
            return_message,
            dry_run,
        )
    update_facilities = []
    create_facilities = {}
    drop_substances = []
    create_substances = []
    update_sources = []
    create_sources = {}
    activitycode_columns = [key for key in df.columns if key.startswith("activitycode")]
    row_nr = 2
    for row_key, row in df.iterrows():
        row_dict = row.to_dict()

        # initialize activitycodes
        source_data = {
            "activitycode1": None,
            "activitycode2": None,
            "activitycode3": None,
        }

        # get pointsource coordinates
        try:
            if pd.isnull(row_dict["lat"]) or pd.isnull(row_dict["lon"]):
                return_message = import_error(
                    f"missing coordinates for source '{row_key}'",
                    return_message,
                    dry_run,
                )
            x = float(row_dict["lat"])
            y = float(row_dict["lon"])
        except ValueError:
            return_message = import_error(
                f"Invalid coordinates on row {row_nr}", return_message, dry_run
            )

        # create geometry
        source_data["geom"] = Point(x, y, srid=srid or project_srid).transform(
            4326, clone=True
        )

        # get chimney properties
        for attr, key in {
            "chimney_height": "chimney_height",
            "chimney_inner_diameter": "inner_diameter",
            "chimney_outer_diameter": "outer_diameter",
            "chimney_gas_speed": "gas_speed",
            "chimney_gas_temperature": "gas_temperature",
        }.items():
            if pd.isna(row_dict[key]):
                return_message = import_error(
                    f"Missing value for {key} on row {row_nr}", return_message, dry_run
                )
            else:
                source_data[attr] = row_dict[key]

        # get downdraft parameters
        if not pd.isnull(row_dict["house_width"]):
            source_data["house_width"] = row_dict["house_width"]
        if not pd.isnull(row_dict["house_height"]):
            source_data["house_height"] = row_dict["house_height"]

        # get activitycodes
        for code_ind, code_set in enumerate(code_sets, 1):
            try:
                code_set_slug = CodeSet.objects.filter(id=code_ind).first().slug
                code_attribute = f"activitycode_{code_set_slug}"
                if code_attribute in row_dict:
                    code = row_dict[code_attribute]
                    if len(code_set) == 0:
                        if code is not None and code is not np.nan:
                            return_message = import_error(
                                f"Unknown activitycode_{code_set_slug} '{code}'"
                                + f" on row {row_nr}",
                                return_message,
                                dry_run,
                            )
                    if not pd.isnull(code):
                        try:
                            # note this can be problematic with codes 01 etc as SNAP
                            # TODO activitycodes should be string directly on import!
                            activity_code = code_set[str(code)]
                            codeset_id = activity_code.code_set_id
                            source_data[f"activitycode{codeset_id}"] = activity_code
                        except KeyError:
                            return_message = import_error(
                                f"Unknown activitycode_{code_set_slug} '{code}'"
                                + f" on row {row_nr}",
                                return_message,
                                dry_run,
                            )
            except AttributeError:
                # no such codeset exists
                if len(activitycode_columns) > len(CodeSet.objects.all()):
                    # need to check if activitycode is specified for unimported codeset
                    codeset_slug = [
                        column.split("_", 1)[-1] for column in activitycode_columns
                    ]
                    for index, column in enumerate(activitycode_columns):
                        if not pd.isnull(row_dict[column]):
                            try:
                                CodeSet.objects.get(slug=codeset_slug[index])
                            except CodeSet.DoesNotExist:
                                return_message = import_error(
                                    f"Specified activitycode {row_dict[column]} for "
                                    + f" unknown codeset {codeset_slug[index]}"
                                    + f" on row {row_nr}",
                                    return_message,
                                    dry_run,
                                )

                pass

        # get columns with tag values for the current row
        tag_keys = [key for key in row_dict.keys() if key.startswith("tag:")]
        # set tags dict for source
        source_data["tags"] = {
            key[4:]: row_dict[key] for key in tag_keys if pd.notna(row_dict[key])
        }

        # get timevar name and corresponding timevar
        timevar_name = row_dict["timevar"]
        if pd.notna(timevar_name):
            try:
                source_data["timevar"] = timevars[timevar_name]
            except KeyError:
                return_message = import_error(
                    f"Timevar '{timevar_name}' " f"on row {row_nr} does not exist",
                    return_message,
                    dry_run,
                )

        # get all column-names starting with "subst" whith value for the current row
        subst_keys = [
            key
            for key in row_dict.keys()
            if key.startswith("subst:") and pd.notna(row_dict[key])
        ]

        # create list of data dict for each substance emission
        emissions = {}
        for subst_key in subst_keys:
            subst = subst_key[6:]
            # dict with substance emission properties (value and substance)
            emis = {}
            emissions[subst] = emis

            # get substance
            try:
                emis["substance"] = substances[subst]
            except KeyError:
                return_message = import_error(
                    f"Undefined substance {subst}", return_message, dry_run
                )

            try:
                if not pd.isnull(row_dict["unit"]):
                    emis["value"] = emission_unit_to_si(
                        float(row_dict[subst_key]), row_dict["unit"]
                    )
                else:
                    return_message = import_error(
                        f"No unit specified for emissions on {row_nr}",
                        return_message,
                        dry_run,
                    )
            except ValueError:
                return_message = import_error(
                    f"Invalid emission value {row_dict[subst_key]} on row {row_nr}",
                    return_message,
                    dry_run,
                )
            except KeyError as err:
                return_message = import_error(f"{err}", return_message, dry_run)

        official_facility_id, source_name = row_key
        if pd.isna(official_facility_id):
            official_facility_id = None

        if pd.isna(source_name):
            return_message = import_error(
                f"No name specified for point-source on row {row_nr}",
                return_message,
                dry_run,
            )

        if pd.isna(row_dict["facility_name"]):
            facility_name = None
        else:
            facility_name = row_dict["facility_name"]

        try:
            facility = facilities[official_facility_id]
            update_facilities.append(facility)
        except KeyError:
            if official_facility_id is not None:
                if official_facility_id in create_facilities:
                    facility = create_facilities[official_facility_id]
                else:
                    facility = Facility(
                        name=facility_name,
                        official_id=official_facility_id,
                    )
                    create_facilities[official_facility_id] = facility
            else:
                facility = None

        source_data["facility"] = facility
        source_key = (official_facility_id, source_name)
        try:
            source = pointsources[source_key]
            for key, val in source_data.items():
                setattr(source, key, val)
            update_sources.append(source)
            drop_substances += list(source.substances.all())
            create_substances += [
                PointSourceSubstance(source=source, **emis)
                for emis in emissions.values()
            ]
        except KeyError:
            source = PointSource(name=source_name, **source_data)
            if source_key not in create_sources:
                create_sources[source_key] = source
                create_substances += [
                    PointSourceSubstance(source=source, **emis)
                    for emis in emissions.values()
                ]
            else:
                return_message = import_error(
                    f"multiple rows for the same point-source '{source_name}'",
                    return_message,
                    dry_run,
                )
        row_nr += 1

    existing_facility_names = set([f.name for f in facilities.values()])
    duplicate_facility_names = []
    for official_id, f in create_facilities.items():
        if f.name in existing_facility_names:
            duplicate_facility_names.append(f.name)
    if len(duplicate_facility_names) > 0:
        return_message = import_error(
            "The following facility names are already used in inventory but "
            f"for facilities with different official_id: {duplicate_facility_names}",
            return_message,
            dry_run,
        )
    duplicate_facility_names = {}
    for f in create_facilities.values():
        if f.name in duplicate_facility_names:
            duplicate_facility_names[f.name] += 1
        else:
            duplicate_facility_names[f.name] = 1
    duplicate_facility_names = [
        name for name, nr in duplicate_facility_names.items() if nr > 1
    ]
    if len(duplicate_facility_names) > 0:
        return_message = import_error(
            "The same facility name is used on multiple rows but "
            f"with different facility_id: {duplicate_facility_names}",
            return_message,
            dry_run,
        )

    if not dry_run:
        Facility.objects.bulk_create(create_facilities.values())
        Facility.objects.bulk_update(update_facilities, ["name"])

    # ensure PointSource.facility_id is not None
    for source in create_sources.values():
        if source.facility is not None:
            # changed by Eef, because IDs were None
            source.facility_id = Facility.objects.get(official_id=source.facility).id

    if not dry_run:
        PointSource.objects.bulk_create(create_sources.values())
        PointSource.objects.bulk_update(
            update_sources,
            [
                "name",
                "geom",
                "tags",
                "chimney_gas_speed",
                "chimney_gas_temperature",
                "chimney_height",
                "chimney_inner_diameter",
                "chimney_outer_diameter",
                "house_height",
                "house_width",
                "activitycode1",
                "activitycode2",
                "activitycode3",
            ],
        )

        # drop existing substance emissions of point-sources that will be updated
        PointSourceSubstance.objects.filter(
            pk__in=[inst.id for inst in drop_substances]
        ).delete()

        # ensure PointSourceSubstance.source_id is not None
        for emis in create_substances:
            emis.source_id = PointSource.objects.get(
                name=emis.source, facility_id=emis.source.facility_id
            ).id
        PointSourceSubstance.objects.bulk_create(create_substances)
    return_dict = {
        "facility": {
            "updated": len(update_facilities),
            "created": len(create_facilities),
        },
        "pointsource": {"updated": len(update_sources), "created": len(create_sources)},
    }
    return return_dict, return_message


def import_timevars(timevar_data, overwrite=False, dry_run=False):
    """import time-variation profiles."""

    # Timevar instances must not be created by bulk_create as the save function
    # is overloaded to calculate the normation constant.
    def make_timevar(data, timevarclass, subname=None, dry_run=False):
        retdict = {}
        return_message = ""
        for name, timevar_data in data.items():
            try:
                typeday = timevar_data["typeday"]
                month = timevar_data["month"]
                if overwrite:
                    newobj, _ = timevarclass.objects.update_or_create(
                        name=name,
                        defaults={"typeday": typeday, "month": month},
                    )
                else:
                    try:
                        newobj = timevarclass.objects.create(
                            name=name, typeday=typeday, month=month
                        )
                    except IntegrityError:
                        raise IntegrityError(
                            f"{timevarclass.__name__} {name} "
                            f"already exist in inventory."
                        )
                retdict[name] = newobj
            except KeyError:
                return_message = import_error(
                    f"Invalid specification of timevar {name}"
                    f", are 'typeday' and 'month' given?",
                    return_message,
                    dry_run,
                )
        return retdict, return_message

    timevars = {}
    for vartype, subdict in timevar_data.items():
        if vartype == "emission":
            timevars["emission"], return_message = make_timevar(
                timevar_data[vartype], Timevar
            )
        else:
            return_message = import_error(
                f"invalid time-variation type '{vartype}' specified",
                return_message,
                dry_run,
            )
    return timevars, return_message


def import_eea_emfacs(filepath, encoding=None):
    """Import point-sources from xlsx or csv-file.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
    """
    existing_eea_emfacs = EEAEmissionFactor.objects.all()
    if len(existing_eea_emfacs) > 1:
        log.debug("emfacs have previously been imported")
        log.debug("for now delete all previous emfacs")
        # TODO if automatically linking EEA emfacs to applied emfacs,
        # then cannot remove but can only update, to avoid removing
        # an emfac which is used for a certain activity.
        existing_eea_emfacs.delete()

    substances = cache_queryset(Substance.objects.all(), "slug")

    extension = filepath.suffix
    if extension == ".csv":
        # read csv-file
        with open(filepath, encoding=encoding or "utf-8") as csvfile:
            log.debug("reading point-sources from csv-file")
            df = pd.read_csv(
                csvfile,
                sep=";",
                skip_blank_lines=True,
                comment="#",
            )
    elif extension == ".xlsx":
        # read spreadsheet
        try:
            workbook = load_workbook(filename=filepath)
        except Exception as exc:
            raise ImportError(str(exc))
        worksheet = workbook.worksheets[0]
        if len(workbook.worksheets) > 1:
            log.debug("debug: multiple sheets in spreadsheet, only importing 1st.")
        data = worksheet.values
        df = worksheet_to_dataframe(data)
        # TODO could replace NA by None?
        # df = df.replace(to_replace="NA", value=None)
    else:
        raise ImportError("Only xlsx and csv files are supported for import")

    row_nr = 2
    no_value_count = 0
    no_unit_count = 0
    create_eea_emfac = []
    for row_key, row in df.iterrows():
        emfac_data = {}
        row_dict = row.to_dict()
        for attr, key in {
            "nfr_code": "NFR",
            "sector": "Sector",
            "table": "Table",
            "tier": "Type",
            "technology": "Technology",
            "fuel": "Fuel",
            "abatement": "Abatement",
            "region": "Region",
            "substance": "Pollutant",
            "value": "Value",
            "unit": "Unit",
            "lower": "CI_lower",
            "upper": "CI_upper",
            "reference": "Reference",
        }.items():
            emfac_data[attr] = row_dict[key]
        # check if substance known
        try:
            subst = emfac_data["substance"]
        except KeyError:
            print(f"No pollutant given, ignoring row {row_nr}")
            row_nr += 1
            no_value_count += 1
            continue
        try:
            emfac_data["substance"] = substances[subst]
        except KeyError:
            if subst == "PM2.5":
                emfac_data["substance"] = substances["PM25"]
            else:
                # TODO log warning
                # many undefined substances in EEA so dont want to raise import warning
                print(f"Undefined substance {subst}")
                print("Saving pollutant as unknown_substance.")
                emfac_data["unknown_substance"] = subst
                emfac_data["substance"] = None
        if row_dict["Value"] is None:
            if (row_dict["CI_lower"] is None) and (row_dict["CI_upper"] is None):
                # TODO log warning
                print(f"No emission factor given, ignoring row {row_nr}")
                row_nr += 1
                no_value_count += 1
                continue
            else:
                # taking mean if both upper and lower not nan, if only one not nan,
                # take that value.
                emfac_data["value"] = np.nanmean(
                    np.array(
                        [row_dict["CI_lower"], row_dict["CI_upper"]], dtype=np.float64
                    )
                )
        if emfac_data["unit"] is None:
            # TODO log warning
            print(f"No unit given, ignoring row {row_nr}")
            row_nr += 1
            no_unit_count += 1
            continue
        # set data in EEA emfac data model
        try:
            float(emfac_data["value"])
        except ValueError:
            print(f"Non numerical value, ignoring row {row_nr}")
            row_nr += 1
            no_value_count += 1
            continue
        eea_emfac = EEAEmissionFactor()
        for key, val in emfac_data.items():
            setattr(eea_emfac, key, val)
        create_eea_emfac.append(eea_emfac)
        row_nr += 1

    EEAEmissionFactor.objects.bulk_create(create_eea_emfac)
    # TODO check for existing emfac and do not create twice, or only update
    return create_eea_emfac


def import_pointsourceactivities(
    filepath,
    encoding=None,
    srid=None,
    import_sheets=SHEET_NAMES,
    return_message="",
    dry_run=False,
):
    """Import point-sources from xlsx or csv-file.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
        srid: srid of file, default is same srid as domain
    """
    try:
        workbook = load_workbook(filename=filepath, data_only=True)
    except Exception as exc:
        return_message = import_error(str(exc), return_message, dry_run)

    return_dict = {}
    sheet_names = [sheet.title for sheet in workbook.worksheets]
    if ("Timevar" in sheet_names) and ("Timevar" in import_sheets):
        timevar_data = workbook["Timevar"].values
        df_timevar = worksheet_to_dataframe(timevar_data)
        timevar_dict = {"emission": {}}
        # NB this only works if Excel file has exact same format
        nr_timevars = (len(df_timevar["ID"]) + 1) // 27
        for i in range(nr_timevars):
            label = df_timevar["ID"][i * 27]
            typeday = np.asarray(
                df_timevar[
                    [
                        "monday",
                        "tuesday",
                        "wednesday",
                        "thursday",
                        "friday",
                        "saturday",
                        "sunday",
                    ]
                ][i * 27 : i * 27 + 24]
            )
            month = np.asarray(df_timevar.iloc[i * 27 + 25, 2:14])
            typeday_str = np.array2string(typeday).replace("\n", "").replace(" ", ", ")
            month_str = np.array2string(month).replace("\n", "").replace(" ", ", ")
            timevar_dict["emission"].update(
                {label: {"typeday": typeday_str, "month": month_str}}
            )
        tv, return_append = import_timevars(
            timevar_dict, overwrite=True, dry_run=dry_run
        )
        return_message += return_append
        return_dict.update({"timevar": {"updated or created": len(tv["emission"])}})

    if ("CodeSet" in sheet_names) and ("CodeSet" in import_sheets):
        nr_codesets = len(CodeSet.objects.all())
        data = workbook["CodeSet"].values
        df_codeset = worksheet_to_dataframe(data)
        slugs = df_codeset["slug"]
        update_codesets = []
        create_codesets = {}
        for row_nr, slug in enumerate(slugs):
            try:
                codeset = CodeSet.objects.get(slug=slug)
                setattr(codeset, "name", df_codeset["name"][row_nr])
                setattr(codeset, "description", df_codeset["description"][row_nr])
                update_codesets.append(codeset)
            except CodeSet.DoesNotExist:
                if nr_codesets < 4:
                    codeset = CodeSet(
                        name=df_codeset["name"][row_nr],
                        slug=slug,
                        description=df_codeset["description"][row_nr],
                    )
                    if slug not in create_codesets:
                        create_codesets[slug] = codeset
                else:
                    return_message = import_error(
                        "Trying to import a new codeset, but can have maximum 3.",
                        return_message,
                        dry_run,
                    )
        if not dry_run:
            CodeSet.objects.bulk_create(create_codesets.values())
            CodeSet.objects.bulk_update(
                update_codesets,
                [
                    "name",
                    "description",
                ],
            )
        return_dict.update(
            {
                "codeset": {
                    "updated": len(update_codesets),
                    "created": len(create_codesets),
                }
            }
        )

    if ("ActivityCode" in sheet_names) and ("ActivityCode" in import_sheets):
        data = workbook["ActivityCode"].values
        df_activitycode = worksheet_to_dataframe(data)
        update_activitycodes = []
        create_activitycodes = {}
        for row_key, row in df_activitycode.iterrows():
            row_dict = row.to_dict()
            try:
                codeset = CodeSet.objects.get(slug=row_dict["codeset_slug"])
            except CodeSet.DoesNotExist:
                return_message = import_error(
                    f"Trying to import an activity code from row '{row_nr}'"
                    + f"but CodeSet '{row_dict['codeset_slug']}' is not defined.",
                    return_message,
                    dry_run,
                )
            if row_dict["vertical_distribution_slug"] is not None:
                try:
                    vdist = VerticalDist.objects.get(
                        slug=row_dict["vertical_distribution_slug"]
                    )
                    vdist_id = vdist.id
                except VerticalDist.DoesNotExist:
                    return_message = import_error(
                        f"Trying to import an activity code from row '{row_nr}'"
                        + "but Vertical Distribution "
                        + f"'{row_dict['vertical_distribution_slug']}'"
                        + " is not defined.",
                        return_message,
                        dry_run,
                    )
            else:
                vdist_id = None
            try:
                activitycode = ActivityCode.objects.get(
                    code_set_id=codeset.id, code=row_dict["activitycode"]
                )
                setattr(activitycode, "label", row_dict["label"])
                setattr(activitycode, "vertical_dist_id", vdist_id)
                update_activitycodes.append(activitycode)
            except ActivityCode.DoesNotExist:
                activitycode = ActivityCode(
                    code=row_dict["activitycode"],
                    label=row_dict["label"],
                    code_set_id=codeset.id,
                    vertical_dist_id=vdist_id,
                )
                create_activitycodes[row_dict["activitycode"]] = activitycode
        if not dry_run:
            ActivityCode.objects.bulk_create(create_activitycodes.values())
            ActivityCode.objects.bulk_update(
                update_activitycodes,
                [
                    "label",
                    "vertical_dist_id",
                ],
            )
        return_dict.update(
            {
                "activitycode": {
                    "updated": len(update_activitycodes),
                    "created": len(create_activitycodes),
                }
            }
        )

    # Could be that activities are linked to previously imported pointsources,
    # or pointsources to be imported later, therefore not requiring PointSource-sheet.
    if ("PointSource" in sheet_names) and ("PointSource" in import_sheets):
        ps, return_message = import_pointsources(
            filepath, srid=srid, return_message=return_message, dry_run=dry_run
        )
        return_dict.update(ps)

    if ("Activity" in sheet_names) and ("Activity" in import_sheets):
        activities = cache_queryset(
            Activity.objects.prefetch_related("emissionfactors").all(), "name"
        )
        data = workbook["Activity"].values
        df_activity = worksheet_to_dataframe(data)
        activity_names = df_activity["activity_name"]
        update_activities = []
        create_activities = {}
        drop_emfacs = []
        for row_nr, activity_name in enumerate(activity_names):
            try:
                activity = activities[activity_name]
                setattr(activity, "name", activity_name)
                setattr(activity, "unit", df_activity["activity_unit"][row_nr])
                update_activities.append(activity)
                drop_emfacs += list(activities[activity_name].emissionfactors.all())
            except KeyError:
                activity = Activity(
                    name=activity_name, unit=df_activity["activity_unit"][row_nr]
                )
                if activity_name not in create_activities:
                    create_activities[activity_name] = activity
                else:
                    return_message = import_error(
                        f"multiple rows for the same activity '{activity_name}'",
                        return_message,
                        dry_run,
                    )
        if not dry_run:
            Activity.objects.bulk_create(create_activities.values())
            Activity.objects.bulk_update(
                update_activities,
                [
                    "name",
                    "unit",
                ],
            )
            # drop existing emfacs of activities that will be updated
            EmissionFactor.objects.filter(
                pk__in=[inst.id for inst in drop_emfacs]
            ).delete()
        return_dict.update(
            {
                "activity": {
                    "updated": len(update_activities),
                    "created": len(create_activities),
                }
            }
        )

    if ("EmissionFactor" in sheet_names) and ("EmissionFactor" in import_sheets):
        data = workbook["EmissionFactor"].values
        df_emfac = worksheet_to_dataframe(data)
        substances = cache_queryset(Substance.objects.all(), "slug")
        activities = cache_queryset(Activity.objects.all(), "name")
        # unique together activity_name and substance
        emissionfactors = cache_queryset(
            EmissionFactor.objects.all(), ["activity", "substance"]
        )
        update_emfacs = []
        create_emfacs = []
        for row_nr in range(len(df_emfac)):
            activity_name = df_emfac.iloc[row_nr]["activity_name"]
            try:
                activity = activities[activity_name]
            except KeyError:
                return_message = import_error(
                    f"unknown activity '{activity_name}'"
                    + f" for emission factor on row '{row_nr}'",
                    return_message,
                    dry_run,
                )
            subst = df_emfac.iloc[row_nr]["substance"]
            try:
                substance = substances[subst]
            except KeyError:
                if subst == "PM2.5":
                    substance = substances["PM25"]
                else:
                    return_message = import_error(
                        f"unknown substance '{subst}'"
                        + f" for emission factor on row '{row_nr}'",
                        return_message,
                        dry_run,
                    )
            factor = df_emfac.iloc[row_nr]["factor"]
            factor_unit = df_emfac.iloc[row_nr]["emissionfactor_unit"]
            activity_quantity_unit, time_unit = activity.unit.split("/")
            mass_unit, factor_quantity_unit = factor_unit.split("/")
            if activity_quantity_unit != factor_quantity_unit:
                # emission factor and activity need to have the same unit for quantity
                # be it GJ, m3 pellets, number of produces bottles, it has to be same
                return_message = import_error(
                    f"Units for emission factor and activity rate for '{activity_name}'"
                    + " are inconsistent, convert units before importing.",
                    return_message,
                    dry_run,
                )
            else:
                factor = activity_ef_unit_to_si(factor, factor_unit)
            try:
                emfac = emissionfactors[(activity, substance)]
                setattr(emfac, "activity", activity)
                setattr(emfac, "substance", substance)
                setattr(emfac, "factor", factor)
                update_emfacs.append(emfac)
            except KeyError:
                emfac = EmissionFactor(
                    activity=activity, substance=substance, factor=factor
                )
                create_emfacs.append(emfac)
        # TODO should check uniquetogether constraint for activity and substance?
        # could be done to give more informative error than integrity error here.
        if not dry_run:
            try:
                EmissionFactor.objects.bulk_create(create_emfacs)
            except IntegrityError:
                return_message = import_error(
                    "Two emission factors for same activity and substance are given.",
                    return_message,
                    dry_run,
                )
            EmissionFactor.objects.bulk_update(
                update_emfacs, ["activity", "substance", "factor"]
            )
        return_dict.update(
            {
                "emission_factors": {
                    "updated": len(update_emfacs),
                    "created": len(create_emfacs),
                }
            }
        )

    if ("PointSource" in sheet_names) and ("PointSource" in import_sheets):
        # now that activities, pointsources and emission factors are created,
        # pointsourceactivities can be created.
        # should not matter whether activities and emission factors were imported from
        # same file or existed already in database.
        pointsourceactivities = cache_queryset(
            PointSourceActivity.objects.all(), ["activity", "source"]
        )
        # TODO check unique activity for source
        data = workbook["PointSource"].values
        df_pointsource = worksheet_to_dataframe(data)
        # TODO add: if any df_pointsource['activity_name'] is not None?
        # otherwise can skip this.
        activities = cache_queryset(Activity.objects.all(), "name")
        pointsources = cache_pointsources(
            PointSource.objects.select_related("facility")
            .prefetch_related("substances")
            .all()
        )
        create_pointsourceactivities = []
        update_pointsourceactivities = []
        for row_key, row in df_pointsource.iterrows():
            if "activity_name" in row:
                if row["activity_name"] is not None:
                    rate = row["activity_rate"]
                    try:
                        activity = activities[row["activity_name"]]
                    except KeyError:
                        return_message = import_error(
                            f"unknown activity '{activity_name}'"
                            + f" for pointsource '{row['source_name']}'",
                            return_message,
                            dry_run,
                        )
                    rate = activity_rate_unit_to_si(rate, activity.unit)
                    # original unit stored in activity.unit, but
                    # pointsourceactivity.rate stored as activity / s.
                    pointsource = pointsources[
                        str(row["facility_id"]), row["source_name"]
                    ]
                    try:
                        psa = pointsourceactivities[activity, pointsource]
                        setattr(psa, "rate", rate)
                        update_pointsourceactivities.append(psa)
                    except KeyError:
                        psa = PointSourceActivity(
                            activity=activity, source=pointsource, rate=rate
                        )
                        create_pointsourceactivities.append(psa)
        if not dry_run:
            PointSourceActivity.objects.bulk_create(create_pointsourceactivities)
            PointSourceActivity.objects.bulk_update(
                update_pointsourceactivities, ["activity", "source", "rate"]
            )
        return_dict.update(
            {
                "pointsourceactivity": {
                    "updated": len(update_pointsourceactivities),
                    "created": len(create_pointsourceactivities),
                }
            }
        )

    return return_dict, return_message

    # TODO
    # add tests
    # figure out how to handle facility, is this really useful?
    # see also discussion about uniqueness wrt facility on mattermost.
    # is it correct that activity rate is not converted to SI?
    # make it easier to validate files for import? or better feedback for correction
    # add code and test for rasterizer that aggregates emissions
