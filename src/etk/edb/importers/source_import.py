"""Data importers for the edb application."""

import logging

import numpy as np
import pandas as pd
from django.contrib.gis.geos import GEOSGeometry, Point
from django.db import IntegrityError
from openpyxl import load_workbook

from etk.edb.cache import cache_queryset
from etk.edb.const import SHEET_NAMES, WGS84_SRID
from etk.edb.models import (
    Activity,
    AreaSource,
    AreaSourceActivity,
    AreaSourceSubstance,
    CodeSet,
    EmissionFactor,
    Facility,
    PointSource,
    PointSourceActivity,
    PointSourceSubstance,
    Substance,
)
from etk.edb.models.timevar_models import Timevar
from etk.edb.units import (
    activity_ef_unit_to_si,
    activity_rate_unit_to_si,
    emission_unit_to_si,
)

from .codeset_import import import_activitycodesheet, import_codesetsheet
from .timevar_import import import_timevarsheet
from .utils import cache_codeset, import_error, worksheet_to_dataframe

# from etk.edb.models.common_models import Settings


# column facility and name are used as index and is therefore not included here
REQUIRED_COLUMNS_AREA = {
    "facility_id": np.str_,
    "facility_name": np.str_,
    "source_name": np.str_,
    "geometry": np.str_,
    "EPSG": int,
    "timevar": np.str_,
}

REQUIRED_COLUMNS_POINT = {
    "facility_id": np.str_,
    "facility_name": np.str_,
    "source_name": np.str_,
    "lat": float,
    "lon": float,
    "timevar": np.str_,
    "chimney_height": float,
    "outer_diameter": float,
    "inner_diameter": float,
    "gas_speed": float,
    "gas_temperature": float,
    # "house_width": float,
    # "house_height": float,
}


log = logging.getLogger(__name__)


def cache_sources(queryset):
    """Return dict of model instances with (facility__official_id, name): instance"""
    sources = {}
    for source in queryset:
        if source.facility is not None:
            sources[source.facility.official_id, source.name] = source
        else:
            sources[None, source.name] = source
    return sources


def import_sources(
    filepath,
    validation=False,
    encoding=None,
    srid=None,
    type="point",
):
    """Import point- or area-sources from xlsx or csv-file.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
        srid: srid of file, default is same srid as domain
    """
    # user defined SRID for import or WGS84 if nothing specified
    # as long as we do not have functions in Eclair to edit the "settings_SRID"
    # it does not make sense to use that SRID as default for import.

    return_message = []
    srid = srid or WGS84_SRID
    # cache related models
    substances = cache_queryset(Substance.objects.all(), "slug")
    timevars = cache_queryset(Timevar.objects.all(), "name")
    facilities = cache_queryset(Facility.objects.all(), "official_id")

    if type == "point":
        sources = cache_sources(
            PointSource.objects.select_related("facility").all()
        )  # .prefetch_related("substances")
    elif type == "area":
        sources = cache_sources(
            AreaSource.objects.select_related("facility").all()
        )  # .prefetch_related("substances")
    else:
        return_message.append(
            import_error(
                "this sourcetype is not implemented",
                validation=validation,
            )
        )

    # using filter.first() here, not get() because code_set{i} does not have to exist
    code_sets = [
        cache_codeset(CodeSet.objects.filter(id=i).first()) for i in range(1, 4)
    ]
    code_set_slugs = {}
    for i in range(1, 4):
        try:
            code_set_slug = CodeSet.objects.get(id=i).slug
        except CodeSet.DoesNotExist:
            code_set_slug = None
        code_set_slugs[i] = code_set_slug

    extension = filepath.suffix
    if extension == ".csv":
        # read csv-file
        if type == "point":
            with open(filepath, encoding=encoding or "utf-8") as csvfile:
                log.debug("reading point-sources from csv-file")
                df = pd.read_csv(
                    csvfile,
                    sep=";",
                    skip_blank_lines=True,
                    comment="#",
                    dtype=REQUIRED_COLUMNS_POINT,
                )
        else:
            with open(filepath, encoding=encoding or "utf-8") as csvfile:
                log.debug("reading area-sources from csv-file")
                df = pd.read_csv(
                    csvfile,
                    sep=";",
                    skip_blank_lines=True,
                    comment="#",
                    dtype=REQUIRED_COLUMNS_AREA,
                )
    elif extension == ".xlsx":
        # read spreadsheet
        try:
            workbook = load_workbook(filename=filepath, data_only=True, read_only=True)
        except Exception as exc:
            return_message = import_error(str(exc), return_message, validation)
        worksheet = workbook.worksheets[0]
        if len(workbook.worksheets) > 1:
            if type == "point":
                log.debug(
                    "Multiple sheets in spreadsheet, importing sheet 'PointSource'."
                )
                data = workbook["PointSource"].values
            elif type == "area":
                log.debug(
                    "Multiple sheets in spreadsheet, importing sheet 'AreaSource'."
                )
                data = workbook["AreaSource"].values
        else:
            data = worksheet.values
        df = worksheet_to_dataframe(data)
        if type == "point":
            df = df.astype(dtype=REQUIRED_COLUMNS_POINT)
        else:
            df = df.astype(dtype=REQUIRED_COLUMNS_AREA)
        # below is necessary not to create facilities with name 'None'
        df = df.replace(to_replace="None", value=None)
        workbook.close()
    else:
        return_message.append(
            import_error(
                "Only xlsx and csv files are supported for import",
                validation=validation,
            )
        )
    if type == "point":
        for col in REQUIRED_COLUMNS_POINT.keys():
            if col not in df.columns:
                return_message.append(
                    import_error(
                        f"Missing required column '{col}'", validation=validation
                    )
                )
    else:
        for col in REQUIRED_COLUMNS_AREA.keys():
            if col not in df.columns:
                return_message.append(
                    import_error(
                        f"Missing required column '{col}'", validation=validation
                    )
                )

    # set dataframe index
    try:
        df.set_index(
            ["facility_id", "source_name"], verify_integrity=True, inplace=True
        )
    except ValueError as err:
        return_message.append(
            import_error(
                f"Non-unique combination of facility_id and source_name: {err}",
                validation=validation,
            )
        )
    update_facilities = []
    create_facilities = {}
    drop_substances = []
    create_substances = []
    update_sources = []
    create_sources = {}
    activitycode_columns = [key for key in df.columns if key.startswith("activitycode")]
    row_nr = 2
    for row_key, row in df.iterrows():
        row_dict = row.to_dict()

        # initialize activitycodes
        source_data = {
            "activitycode1": None,
            "activitycode2": None,
            "activitycode3": None,
        }

        if type == "point":
            # get pointsource coordinates
            try:
                if pd.isnull(row_dict["lat"]) or pd.isnull(row_dict["lon"]):
                    return_message.append(
                        import_error(
                            f"missing coordinates for source '{row_key}'",
                            validation=validation,
                        )
                    )
                x = float(row_dict["lon"])
                y = float(row_dict["lat"])
            except ValueError:
                return_message.append(
                    import_error(
                        f"Invalid {type} coordinates on row {row_nr}",
                        validation=validation,
                    )
                )
            # create geometry
            source_data["geom"] = Point(x, y, srid=srid).transform(4326, clone=True)
            # get chimney properties
            for attr, key in {
                "chimney_height": "chimney_height",
                "chimney_inner_diameter": "inner_diameter",
                "chimney_outer_diameter": "outer_diameter",
                "chimney_gas_speed": "gas_speed",
                "chimney_gas_temperature": "gas_temperature",
            }.items():
                if pd.isna(row_dict[key]):
                    return_message.append(
                        import_error(
                            "Missing value in PointSource sheet "
                            f"for {key} on row {row_nr}",
                            validation=validation,
                        )
                    )
                else:
                    source_data[attr] = row_dict[key]

            # get downdraft parameters
            try:
                if not pd.isnull(row_dict["house_width"]):
                    source_data["house_width"] = row_dict["house_width"]
            except KeyError:
                if row_nr == 2:
                    log.debug("house_width is skipped from import.")
            try:
                if not pd.isnull(row_dict["house_height"]):
                    source_data["house_height"] = row_dict["house_height"]
            except KeyError:
                if row_nr == 2:
                    log.debug("house_heigth is skipped from import.")

        elif type == "area":
            try:
                if pd.isnull(row_dict["geometry"]):
                    return_message.append(
                        import_error(
                            f"missing area polygon for source '{row_key}'",
                            validation,
                        )
                    )
                wkt_polygon = row_dict["geometry"]
                # TODO add check that valid WKT polygon
            except ValueError:
                return_message.append(
                    import_error(
                        f"Invalid polygon geometry in AreaSource sheet on row {row_nr}",
                        validation=validation,
                    )
                )
            # create geometry
            EPSG = row_dict["EPSG"]
            if pd.isnull(EPSG):
                EPSG = 4326
            source_data["geom"] = GEOSGeometry(f"SRID={int(EPSG)};" + wkt_polygon)
        else:
            return_message.append(
                import_error(
                    "this sourcetype is not implemented", validation=validation
                )
            )

        # get activitycodes
        for code_ind, code_set in enumerate(code_sets, 1):
            try:
                code_set_slug = code_set_slugs[code_ind]
                code_attribute = f"activitycode_{code_set_slug}"
                if code_attribute in row_dict:
                    code = row_dict[code_attribute]
                    if len(code_set) == 0:
                        if code is not None and code is not np.nan:
                            return_message.append(
                                import_error(
                                    f"Unknown activitycode_{code_set_slug} '{code}'"
                                    f" for {type} source on row {row_nr}",
                                    validation=validation,
                                )
                            )
                    if not pd.isnull(code):
                        try:
                            # note this can be problematic with codes 01 etc as SNAP
                            # TODO activitycodes should be string directly on import!
                            activity_code = code_set[str(code)]
                            codeset_id = activity_code.code_set_id
                            source_data[f"activitycode{codeset_id}"] = activity_code
                        except KeyError:
                            return_message.append(
                                import_error(
                                    f"Unknown activitycode_{code_set_slug} '{code}'"
                                    f" for {type} source on row {row_nr}",
                                    validation=validation,
                                )
                            )
            except AttributeError:
                # no such codeset exists
                if len(activitycode_columns) > len(CodeSet.objects.all()):
                    # need to check if activitycode is specified for unimported codeset
                    codeset_slug = [
                        column.split("_", 1)[-1] for column in activitycode_columns
                    ]
                    for index, column in enumerate(activitycode_columns):
                        if not pd.isnull(row_dict[column]):
                            try:
                                CodeSet.objects.get(slug=codeset_slug[index])
                            except CodeSet.DoesNotExist:
                                return_message.append(
                                    import_error(
                                        "Specified activitycode "
                                        f"{row_dict[column]} for "
                                        f" unknown codeset {codeset_slug[index]}"
                                        f" for {type} source on row {row_nr}",
                                        validation=validation,
                                    )
                                )
                pass

        # get columns with tag values for the current row
        tag_keys = [key for key in row_dict.keys() if key.startswith("tag:")]
        # set tags dict for source
        source_data["tags"] = {
            key[4:]: row_dict[key] for key in tag_keys if pd.notna(row_dict[key])
        }

        # get timevar name and corresponding timevar
        timevar_name = row_dict["timevar"]
        if pd.notna(timevar_name):
            try:
                source_data["timevar"] = timevars[timevar_name]
            except KeyError:
                return_message.append(
                    import_error(
                        f"Timevar '{timevar_name}' "
                        f"on row {row_nr} for {type} source does not exist",
                        validation=validation,
                    )
                )
        # get all column-names starting with "subst" whith value for the current row
        subst_keys = [
            key
            for key in row_dict.keys()
            if key.startswith("subst:") and pd.notna(row_dict[key])
        ]

        # create list of data dict for each substance emission
        emissions = {}
        for subst_key in subst_keys:
            subst = subst_key[6:]
            # dict with substance emission properties (value and substance)
            emis = {}
            emissions[subst] = emis

            # get substance
            try:
                emis["substance"] = substances[subst]
            except KeyError:
                return_message = import_error(
                    f"Undefined substance {subst}", return_message, validation
                )

            try:
                if not pd.isnull(row_dict["unit"]):
                    emis["value"] = emission_unit_to_si(
                        float(row_dict[subst_key]), row_dict["unit"]
                    )
                else:
                    return_message.append(
                        import_error(
                            f"No unit specified for {type} emissions on row {row_nr}",
                            validation=validation,
                        )
                    )
            except ValueError:
                return_message.append(
                    import_error(
                        f"Invalid {type} emission value {row_dict[subst_key]}"
                        f" on row {row_nr}",
                        validation=validation,
                    )
                )
            except KeyError as err:
                return_message.append(import_error(f"{err}", validation=validation))

        official_facility_id, source_name = row_key
        if pd.isna(official_facility_id):
            official_facility_id = None

        if pd.isna(source_name):
            return_message.append(
                import_error(
                    f"No name specified for {type} source on row {row_nr}",
                    validation=validation,
                )
            )
        if pd.isna(row_dict["facility_name"]):
            facility_name = None
        else:
            facility_name = row_dict["facility_name"]

        try:
            facility = facilities[official_facility_id]
            update_facilities.append(facility)
        except KeyError:
            if official_facility_id is not None:
                if official_facility_id in create_facilities:
                    facility = create_facilities[official_facility_id]
                else:
                    if facility_name is None:
                        return_message.append(
                            import_error(
                                f"No name specified for facility on row {row_nr}",
                                validation=validation,
                            )
                        )
                        facility_name = "unspecified"

                    facility = Facility(
                        name=facility_name,
                        official_id=official_facility_id,
                    )
                    create_facilities[official_facility_id] = facility
            else:
                facility = None

        source_data["facility"] = facility
        source_key = (official_facility_id, source_name)
        try:
            source = sources[source_key]
            for key, val in source_data.items():
                setattr(source, key, val)
            update_sources.append(source)
            drop_substances += list(source.substances.all())
            if type == "point":
                create_substances += [
                    PointSourceSubstance(source=source, **emis)
                    for emis in emissions.values()
                ]
            elif type == "area":
                create_substances += [
                    AreaSourceSubstance(source=source, **emis)
                    for emis in emissions.values()
                ]
        except KeyError:
            if type == "point":
                source = PointSource(name=source_name, **source_data)
                if source_key not in create_sources:
                    create_sources[source_key] = source
                    create_substances += [
                        PointSourceSubstance(source=source, **emis)
                        for emis in emissions.values()
                    ]
                else:
                    return_message.append(
                        import_error(
                            f"multiple rows for the same point-source '{source_name}'",
                            validation=validation,
                        )
                    )
            else:
                source = AreaSource(name=source_name, **source_data)
                if source_key not in create_sources:
                    create_sources[source_key] = source
                    create_substances += [
                        AreaSourceSubstance(source=source, **emis)
                        for emis in emissions.values()
                    ]
                else:
                    return_message.append(
                        import_error(
                            f"multiple rows for the same area-source '{source_name}'",
                            validation=validation,
                        )
                    )
        row_nr += 1

    existing_facility_names = set([f.name for f in facilities.values()])
    duplicate_facility_names = []
    for official_id, f in create_facilities.items():
        if f.name in existing_facility_names:
            duplicate_facility_names.append(f.name)
    if len(duplicate_facility_names) > 0:
        return_message.append(
            import_error(
                "The following facility names are already "
                "used in inventory but for facilities with "
                f"different official_id: {duplicate_facility_names}",
                validation=validation,
            )
        )
    duplicate_facility_names = {}
    for f in create_facilities.values():
        if f.name in duplicate_facility_names:
            duplicate_facility_names[f.name] += 1
        else:
            duplicate_facility_names[f.name] = 1
    duplicate_facility_names = [
        name for name, nr in duplicate_facility_names.items() if nr > 1
    ]
    if len(duplicate_facility_names) > 0:
        return_message.append(
            import_error(
                "The same facility name is used on multiple rows but "
                f"with different facility_id: {duplicate_facility_names}",
                validation=validation,
            )
        )

    Facility.objects.bulk_create(create_facilities.values())
    Facility.objects.bulk_update(update_facilities, ["name"])

    facilities = cache_queryset(Facility.objects.all(), "id")
    # ensure PointSource.facility_id is not None if facility exists.
    for source in create_sources.values():
        if source.facility is not None:
            # find the facility_id corresponding to official id, or set None
            source.facility_id = next(
                (
                    key
                    for key, value in facilities.items()
                    if str(value) == str(source.facility)
                ),
                None,
            )
            if source.facility_id is None:
                raise ImportError(
                    f"Could not link pointsource {source.name} to "
                    + f"facility {source.facility.name}"
                )

    if type == "point":
        PointSource.objects.bulk_create(create_sources.values())
        PointSource.objects.bulk_update(
            update_sources,
            [
                "name",
                "geom",
                "tags",
                "chimney_gas_speed",
                "chimney_gas_temperature",
                "chimney_height",
                "chimney_inner_diameter",
                "chimney_outer_diameter",
                "house_height",
                "house_width",
                "activitycode1",
                "activitycode2",
                "activitycode3",
            ],
        )

        # drop existing substance emissions of point-sources that will be updated
        PointSourceSubstance.objects.filter(
            pk__in=[inst.id for inst in drop_substances]
        ).delete()

        # ensure PointSourceSubstance.source_id is not None
        for emis in create_substances:
            emis.source_id = PointSource.objects.get(
                name=emis.source, facility_id=emis.source.facility_id
            ).id
        PointSourceSubstance.objects.bulk_create(create_substances)
        return_dict = {
            "facility": {
                "updated": len(update_facilities),
                "created": len(create_facilities),
            },
            "pointsource": {
                "updated": len(update_sources),
                "created": len(create_sources),
            },
        }
    if type == "area":
        AreaSource.objects.bulk_create(create_sources.values())
        AreaSource.objects.bulk_update(
            update_sources,
            [
                "name",
                "geom",
                "tags",
                "activitycode1",
                "activitycode2",
                "activitycode3",
            ],
        )

        # drop existing substance emissions of point-sources that will be updated
        AreaSourceSubstance.objects.filter(
            pk__in=[inst.id for inst in drop_substances]
        ).delete()

        # ensure PointSourceSubstance.source_id is not None
        for emis in create_substances:
            emis.source_id = AreaSource.objects.get(
                name=emis.source, facility_id=emis.source.facility_id
            ).id
        AreaSourceSubstance.objects.bulk_create(create_substances)
        return_dict = {
            "facility": {
                "updated": len(update_facilities),
                "created": len(create_facilities),
            },
            "areasource": {
                "updated": len(update_sources),
                "created": len(create_sources),
            },
        }
    return return_dict, return_message


def import_sourceactivities(
    filepath,
    encoding=None,
    srid=None,
    import_sheets=SHEET_NAMES,
    validation=False,
):
    """Import point-sources and/or area-sources from xlsx or csv-file.

    args
        filepath: path to file

    options
        encoding: encoding of file (default is utf-8)
        srid: srid of file, default is same srid as domain
    """
    return_message = []
    try:
        workbook = load_workbook(filename=filepath, data_only=True, read_only=True)
    except Exception as exc:
        return_message = import_error(str(exc), return_message, validation)

    return_dict = {}
    sheet_names = [sheet.title for sheet in workbook.worksheets]
    if ("Timevar" in sheet_names) and ("Timevar" in import_sheets):
        updates, msgs = import_timevarsheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs

    if ("CodeSet" in sheet_names) and ("CodeSet" in import_sheets):
        updates, msgs = import_codesetsheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs

    if ("ActivityCode" in sheet_names) and ("ActivityCode" in import_sheets):
        updates, msgs = import_activitycodesheet(workbook, validation)
        return_dict.update(updates)
        return_message += msgs
    # Could be that activities are linked to previously imported pointsources,
    # or pointsources to be imported later, therefore not requiring PointSource-sheet.
    if ("PointSource" in sheet_names) and ("PointSource" in import_sheets):
        ps, msgs = import_sources(
            filepath,
            srid=srid,
            validation=validation,
            type="point",
        )
        return_dict.update(ps)
        return_message += msgs

    if ("AreaSource" in sheet_names) and ("AreaSource" in import_sheets):
        ps, msgs = import_sources(
            filepath,
            srid=srid,
            validation=validation,
            type="area",
        )
        return_message += msgs
        return_dict.update(ps)

    if ("Activity" in sheet_names) and ("Activity" in import_sheets):
        activities = cache_queryset(
            Activity.objects.prefetch_related("emissionfactors").all(), "name"
        )
        data = workbook["Activity"].values
        df_activity = worksheet_to_dataframe(data)
        activity_names = df_activity["activity_name"]
        update_activities = []
        create_activities = {}
        drop_emfacs = []
        for row_nr, activity_name in enumerate(activity_names):
            try:
                activity = activities[activity_name]
                setattr(activity, "name", activity_name)
                setattr(activity, "unit", df_activity["activity_unit"][row_nr])
                update_activities.append(activity)
                drop_emfacs += list(activities[activity_name].emissionfactors.all())
            except KeyError:
                activity = Activity(
                    name=activity_name, unit=df_activity["activity_unit"][row_nr]
                )
                if activity_name not in create_activities:
                    create_activities[activity_name] = activity
                else:
                    return_message.append(
                        import_error(
                            f"multiple rows for the same activity '{activity_name}'",
                            validation=validation,
                        )
                    )
        Activity.objects.bulk_create(create_activities.values())
        Activity.objects.bulk_update(
            update_activities,
            [
                "name",
                "unit",
            ],
        )
        # drop existing emfacs of activities that will be updated
        EmissionFactor.objects.filter(pk__in=[inst.id for inst in drop_emfacs]).delete()
        return_dict.update(
            {
                "activity": {
                    "updated": len(update_activities),
                    "created": len(create_activities),
                }
            }
        )

    if ("EmissionFactor" in sheet_names) and ("EmissionFactor" in import_sheets):
        data = workbook["EmissionFactor"].values
        df_emfac = worksheet_to_dataframe(data)
        substances = cache_queryset(Substance.objects.all(), "slug")
        activities = cache_queryset(Activity.objects.all(), "name")
        # unique together activity_name and substance
        emissionfactors = cache_queryset(
            EmissionFactor.objects.all(), ["activity", "substance"]
        )
        update_emfacs = []
        create_emfacs = []
        for row_nr in range(len(df_emfac)):
            activity_name = df_emfac.iloc[row_nr]["activity_name"]
            try:
                activity = activities[activity_name]
                subst = df_emfac.iloc[row_nr]["substance"]
                try:
                    substance = substances[subst]
                    factor = df_emfac.iloc[row_nr]["factor"]
                    factor_unit = df_emfac.iloc[row_nr]["emissionfactor_unit"]
                    activity_quantity_unit, time_unit = activity.unit.split("/")
                    mass_unit, factor_quantity_unit = factor_unit.split("/")
                    if activity_quantity_unit != factor_quantity_unit:
                        # emission factor and activity need to have the same unit
                        # for quantity, eg GJ, m3 "pellets", number of produces bottles
                        return_message.append(
                            import_error(
                                "Units for emission factor and activity rate for"
                                f" '{activity_name}'"
                                " are inconsistent, convert units before importing.",
                                validation=validation,
                            )
                        )
                    else:
                        factor = activity_ef_unit_to_si(factor, factor_unit)
                    try:
                        emfac = emissionfactors[(activity, substance)]
                        setattr(emfac, "activity", activity)
                        setattr(emfac, "substance", substance)
                        setattr(emfac, "factor", factor)
                        update_emfacs.append(emfac)
                    except KeyError:
                        emfac = EmissionFactor(
                            activity=activity, substance=substance, factor=factor
                        )
                        create_emfacs.append(emfac)
                except KeyError:
                    if subst == "PM2.5":
                        substance = substances["PM25"]
                    else:
                        return_message.append(
                            import_error(
                                f"unknown substance '{subst}'"
                                f" for emission factor on row '{row_nr}'",
                                validation=validation,
                            )
                        )
            except KeyError:
                return_message.append(
                    import_error(
                        f"unknown activity '{activity_name}'"
                        f" for emission factor on row '{row_nr}'",
                        validation=validation,
                    )
                )

        try:
            EmissionFactor.objects.bulk_create(create_emfacs)
        except IntegrityError:
            return_message.append(
                import_error(
                    "Two emission factors for same activity and substance are given.",
                    validation=validation,
                )
            )
        EmissionFactor.objects.bulk_update(
            update_emfacs, ["activity", "substance", "factor"]
        )
        return_dict.update(
            {
                "emission_factors": {
                    "updated": len(update_emfacs),
                    "created": len(create_emfacs),
                }
            }
        )

    if ("PointSource" in sheet_names) and ("PointSource" in import_sheets):
        # now that activities, pointsources and emission factors are created,
        # pointsourceactivities can be created.
        # should not matter whether activities and emission factors were imported from
        # same file or existed already in database.
        pointsourceactivities = cache_queryset(
            PointSourceActivity.objects.select_related("activity", "source").all(),
            ["activity", "source"],
        )
        # TODO check unique activity for source
        data = workbook["PointSource"].values
        df_pointsource = worksheet_to_dataframe(data)
        activities = cache_queryset(Activity.objects.all(), "name")
        pointsources = cache_sources(
            PointSource.objects.select_related("facility").all()
        )
        create_pointsourceactivities = []
        update_pointsourceactivities = []
        for row_key, row in df_pointsource.iterrows():
            # NB: does not work if column header starts with space, but same for subst:
            activity_keys = [k for k in row.keys() if k.startswith("act:")]
            for activity_key in activity_keys:
                if not pd.isnull(row[activity_key]):
                    rate = row[activity_key]
                    try:
                        activity = activities[activity_key[4:]]
                    except KeyError:

                        return_message.append(
                            import_error(
                                f"unknown activity '{activity_name}'"
                                + f" for pointsource '{row['source_name']}'",
                                validation=validation,
                            )
                        )
                    rate = activity_rate_unit_to_si(rate, activity.unit)
                    # original unit stored in activity.unit, but
                    # pointsourceactivity.rate stored as activity / s.
                    pointsource = pointsources[
                        str(row["facility_id"]), str(row["source_name"])
                    ]
                    try:
                        psa = pointsourceactivities[activity, pointsource]
                        setattr(psa, "rate", rate)
                        update_pointsourceactivities.append(psa)
                    except KeyError:
                        psa = PointSourceActivity(
                            activity=activity, source=pointsource, rate=rate
                        )
                        create_pointsourceactivities.append(psa)

        PointSourceActivity.objects.bulk_create(create_pointsourceactivities)
        PointSourceActivity.objects.bulk_update(
            update_pointsourceactivities, ["activity", "source", "rate"]
        )
        return_dict.update(
            {
                "pointsourceactivity": {
                    "updated": len(update_pointsourceactivities),
                    "created": len(create_pointsourceactivities),
                }
            }
        )

    if ("AreaSource" in sheet_names) and ("AreaSource" in import_sheets):
        areasourceactivities = cache_queryset(
            AreaSourceActivity.objects.select_related("activity", "source").all(),
            ["activity", "source"],
        )
        # TODO check unique activity for source
        data = workbook["AreaSource"].values
        df_areasource = worksheet_to_dataframe(data)
        activities = cache_queryset(Activity.objects.all(), "name")
        areasources = cache_sources(
            AreaSource.objects.select_related("facility")
            .prefetch_related("substances")
            .all()
        )
        create_areasourceactivities = []
        update_areasourceactivities = []
        for row_key, row in df_areasource.iterrows():
            # NB: does not work if column header starts with space, but same for subst:
            activity_keys = [k for k in row.keys() if k.startswith("act:")]
            for activity_key in activity_keys:
                if not pd.isnull(row[activity_key]):
                    rate = float(row[activity_key])
                    try:
                        activity = activities[activity_key[4:]]
                    except KeyError:
                        return_message.append(
                            import_error(
                                f"unknown activity '{activity_name}'"
                                f" for areasource '{row['source_name']}'",
                                validation=validation,
                            )
                        )
                    rate = activity_rate_unit_to_si(rate, activity.unit)
                    # original unit stored in activity.unit, but
                    # areasourceactivity.rate stored as activity / s.
                    areasource = areasources[
                        str(row["facility_id"]), str(row["source_name"])
                    ]
                    try:
                        psa = areasourceactivities[activity, areasource]
                        setattr(psa, "rate", rate)
                        update_areasourceactivities.append(psa)
                    except KeyError:
                        psa = AreaSourceActivity(
                            activity=activity, source=areasource, rate=rate
                        )
                        create_areasourceactivities.append(psa)

        AreaSourceActivity.objects.bulk_create(create_areasourceactivities)
        AreaSourceActivity.objects.bulk_update(
            update_areasourceactivities, ["activity", "source", "rate"]
        )
        return_dict.update(
            {
                "areasourceactivity": {
                    "updated": len(update_areasourceactivities),
                    "created": len(create_areasourceactivities),
                }
            }
        )

    workbook.close()
    return return_dict, return_message
